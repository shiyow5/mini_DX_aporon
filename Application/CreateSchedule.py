import pandas as pd
import numpy as np
import os
import json
import uuid
import unicodedata
from pathlib import Path
import datetime
from datetime import timedelta
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side
from Excel2Dataframe import get_deadline


class Scheduling:
    def __init__(self, order_data_path: str, ref_data_path: str = None, save_data_path: str = None):
        ROOT_PATH = Path(__file__).resolve().parent.parent
        
        self.order_data_path: str = order_data_path
        
        if ref_data_path:
            self.ref_data_path: str = ref_data_path
        else:
            self.ref_data_path: str = os.path.join(ROOT_PATH, "Datas/production_create_description.json")
        
        if save_data_path:
            self.save_data_path: str = save_data_path
        else:
            filename = os.path.basename(self.order_data_path).replace(".xlsx", "")
            self.save_data_path: str = os.path.join(ROOT_PATH, f"{filename}_schedule_{uuid.uuid4()}.xlsx")
        
        try:
            self.order_data: pd.DataFrame = get_deadline(self.order_data_path)
        except Exception as e:
            raise Exception(f"注文データ抽出中にエラーが発生\n{e}")
        
        try:
            with open(self.ref_data_path, "r", encoding="utf-8") as f:
                self.ref_data: dict = json.load(f)
        except Exception as e:
            raise Exception(f"名称2参照データ抽出中にエラーが発生\n{e}")
    
    def order2baseschedule(self):
        required_columns = {"製品規格名称", "名称2"}
        if not required_columns.issubset(self.order_data.columns):
            raise ValueError(f"Missing required columns: {required_columns - set(self.order_data.columns)}")
    
        date_columns = [col for col in self.order_data.columns if (col not in required_columns) and isinstance(col, datetime.date)]
        
        transformed_rows = []
        miss_datas = []
    
        for _, row in self.order_data.iterrows():
            base_row = {"製品規格名称": row["製品規格名称"], "名称2": row["名称2"], "工程名": None}
        
            if row["名称2"] in self.ref_data.keys():
                processing = self.ref_data[row["名称2"]]["processing"]
                new_rows = [{**base_row, **{col: np.nan for col in date_columns}} for _ in range(len(processing))]
            else:
                miss_datas.append((row["製品規格名称"], row["名称2"]))
                processing = None
                new_rows = None
        
            for date_col in date_columns:
                value = row[date_col]
                if pd.notna(value) and isinstance(value, (int, float)):
                    if new_rows and processing:
                        for i, process_row in enumerate(processing):
                            new_rows[i]["工程名"] = list(self.ref_data[row["名称2"]]["description"].keys())[i]
                            #if i!=0:
                            #    new_rows[i]["製品規格名称"] = None
                            #    new_rows[i]["名称2"] = None
                            for j in range(len(process_row)):
                                now_day = date_col - timedelta(days=len(process_row)-j)
                                new_rows[i][now_day] = process_row[j] * value if process_row[j]!=0 else None
                        
            transformed_rows += new_rows
        
        df = pd.DataFrame(transformed_rows)
        df.to_excel(self.save_data_path, index=False)
    
        return miss_datas
    
    def coloring(self):
        try:
            wb = xl.load_workbook(filename=self.save_data_path)
            ws = wb.worksheets[0]
        except Exception as e:
            raise Exception(f"スケジューリングが実行されていません\n{e}")
        
        thin_border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"), 
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )
        
        for i, row in enumerate(ws, start=1):
            if i==1:
                continue
            color = self.ref_data[ws.cell(row = i, column = 2).value]["color"]
            fill = PatternFill(patternType='solid', fgColor=color)
            
            for cell in row:
                if isinstance(ws[cell.coordinate].value, (int, float)):
                    ws[cell.coordinate].fill = fill
                    ws[cell.coordinate].border = thin_border
                    
        wb.save(self.save_data_path)
    
    def adjust_cell_width(self):
        def calc_width(text):
            """ 全角文字と半角文字を考慮した幅を計算 """
            width = 0
            for char in text:
                if unicodedata.east_asian_width(char) in ['F', 'W']:  # 全角
                    width += 2
                else:  # 半角
                    width += 1
            return width
        
        try:
            wb = xl.load_workbook(filename=self.save_data_path)
            ws = wb.worksheets[0]
        except Exception as e:
            raise Exception(f"スケジューリングが実行されていません\n{e}")
        
        # 各列の最大文字数を取得し、列幅を設定
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # 列のアルファベット（A, B, C, ...）

            for cell in col:
                if cell.value:
                    max_length = max(max_length, calc_width(str(cell.value).replace(" 00:00:00", "")))

            # 列幅を設定
            ws.column_dimensions[col_letter].width = max_length * 1.5

        wb.save(self.save_data_path)
    
    def create(self):
        miss_datas = self.order2baseschedule()
        self.adjust_cell_width()
        self.coloring()
        
        return miss_datas


def test_usecase1():
    test_data = r"C:\Users\Owner\Downloads\学内アルバイト_ミニDX化\mini_DX_aporon\Datas\test_data.xlsx"
    ps = Scheduling(order_data_path=test_data)
    ps.order2baseschedule()
    print("Excelファイルに保存しました。")

def test_usecase2():
    test_data = r"C:\Users\Owner\Downloads\学内アルバイト_ミニDX化\mini_DX_aporon\Datas\test_data.xlsx"
    ps = Scheduling(order_data_path=test_data)
    miss_datas = ps.create()
    if miss_datas:
        for miss_data in miss_datas:
            print(f"製品名：{miss_data[0]}について、\n名称2：{miss_data[1]}が見つかりませんでした。")

test_usecase2()