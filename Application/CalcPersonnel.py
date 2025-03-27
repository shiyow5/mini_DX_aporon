import pandas as pd
import numpy as np
import os
import re
import json
import datetime
import unicodedata
from pathlib import Path
import openpyxl as xl
from openpyxl.styles import PatternFill, Border, Side
from Application.Excel2Dataframe import clean_name

class Calculation:
    def __init__(self, schedule_data_path: str, ref_limit_path: str = None, ref_product_path: str = None, sheet_name: str = None):
        ROOT_PATH = Path(__file__).resolve().parent.parent
        
        self.schedule_data_path: str = schedule_data_path
        
        if ref_limit_path:
            self.ref_limit_path: str = ref_limit_path
        else:
            self.ref_limit_path: str = os.path.join(ROOT_PATH, "Datas/personnel_limit.json")
        
        if ref_product_path:
            self.ref_product_path: str = ref_product_path
        else:
            self.ref_product_path: str = os.path.join(ROOT_PATH, "Datas/production_create_description.json")
        
        if sheet_name:
            self.sheet_name = sheet_name
        else:
            self.sheet_name = "Sheet1"
        
        try:
            self.schedule_data: pd.DataFrame = pd.read_excel(self.schedule_data_path, sheet_name=self.sheet_name)
            self.schedule_data.columns = [col.date() if isinstance(col, datetime.datetime) else col for col in self.schedule_data.columns]
        except Exception as e:
            raise Exception(f"生産計画データ抽出中にエラーが発生\n{e}")
        
        try:
            with open(self.ref_limit_path, "r", encoding="utf-8") as f:
                self.ref_limit: dict = json.load(f)
        except Exception as e:
            raise Exception(f"一人当たりの日産数データ抽出中にエラーが発生\n{e}")
        
        try:
            with open(self.ref_product_path, "r", encoding="utf-8") as f:
                self.ref_product: dict = json.load(f)
        except Exception as e:
            raise Exception(f"名称2参照データ抽出中にエラーが発生\n{e}")
    
    def basecalc(self):
        required_columns = {"製品規格名称", "工程名"}
        if not required_columns.issubset(self.schedule_data.columns):
            raise ValueError(f"Missing required columns: {required_columns - set(self.schedule_data.columns)}")
        
        date_columns = [col for col in self.schedule_data.columns if isinstance(col, datetime.date)]
        
        required_gaikan = pd.DataFrame([[0.0] * len(date_columns)], columns=date_columns)
        required_other = pd.DataFrame([[0.0] * len(date_columns)], columns=date_columns)
        miss_datas = []
        for _, row in self.schedule_data.iterrows():
            process_name = clean_name(row["工程名"])
            ref_targets = {key: value for key, value in self.ref_limit.items() if clean_name(key).startswith(process_name)}
            if not ref_targets:
                miss_datas.append((row["製品規格名称"], row["工程名"]))
            
            if len(ref_targets) >= 2:
                temp_targets = None
                for key, value in ref_targets.items():
                    product_match = re.search(r"\((.*?)\)", key)
                    product_name = product_match.group(1) if product_match else None
                    if product_name and row["製品規格名称"].startswith(product_name):
                        temp_targets = {key: value}
                if temp_targets:
                    ref_targets = temp_targets
                else:
                    ref_targets = {process_name: self.ref_limit.get(process_name, None)}
            
            limit: float = list(ref_targets.values())[0] if ref_targets else None
            
            data_for_calc = row[date_columns].astype("float").fillna(0)
            if "外観検査" in process_name:
                required_gaikan += (data_for_calc / limit) if limit else 0
            else:
                required_other += (data_for_calc / limit) if limit else 0
        
        required_gaikan = required_gaikan.map(lambda num: float(f"{num:.1f}"))
        required_other = required_other.map(lambda num: float(f"{num:.1f}"))
        
        required_gaikan["製品規格名称"] = "外観検査　必要人数"
        required_other["製品規格名称"] = "他　必要人数"
        
        space_row = pd.DataFrame(np.nan, index=[0], columns=date_columns)
        result_df = pd.concat([self.schedule_data, space_row, required_gaikan, required_other], ignore_index=True)
        result_df.to_excel(self.schedule_data_path, index=False)
        
        return miss_datas
    
    def coloring(self):
        try:
            wb = xl.load_workbook(filename=self.schedule_data_path)
            ws = wb[self.sheet_name]
        except Exception as e:
            raise Exception(f"Excelファイルを開けません\n{e}")
        
        thin_border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"), 
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )
        
        for i, row in enumerate(ws, start=1):
            if i==1:
                continue
            ref_description = self.ref_product.get(ws.cell(row = i, column = 2).value, None)
            color = ref_description["color"] if ref_description else "c0c0c0"
            fill = PatternFill(patternType='solid', fgColor=color)
            
            for cell in row:
                if isinstance(ws[cell.coordinate].value, (int, float)):
                    ws[cell.coordinate].fill = fill
                    ws[cell.coordinate].border = thin_border
                    
        wb.save(self.schedule_data_path)
    
    def adjust_cell_width(self):
        def calc_width(text):
            """ 全角文字と半角文字を考慮した幅を計算 """
            width = 0
            for char in text:
                if unicodedata.east_asian_width(char) in ['F', 'W']:  # 全角
                    width += 2
                else:  # 半角
                    width += 1
            return width
        
        try:
            wb = xl.load_workbook(filename=self.schedule_data_path)
            ws = wb[self.sheet_name]
        except Exception as e:
            raise Exception(f"Excelファイルを開けません\n{e}")
        
        # 各列の最大文字数を取得し、列幅を設定
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # 列のアルファベット（A, B, C, ...）

            for cell in col:
                if not isinstance(cell.value, (int, float)):
                    max_length = max(max_length, calc_width(str(cell.value).replace(" 00:00:00", "")))

            # 列幅を設定
            ws.column_dimensions[col_letter].width = max_length * 1.5

        wb.save(self.schedule_data_path)
    
    def reset_calc(self):
        df = self.schedule_data
        delete_row = (df["製品規格名称"]=="外観検査　必要人数") | (df["製品規格名称"]=="他　必要人数")
        delete_row |= (df["製品規格名称"]=="欠勤") | (df["製品規格名称"]=="サンワ") | pd.isna(df["製品規格名称"]) | pd.isna(df["工程名"])
        delete_row |= (df["製品規格名称"]=="合計") | (df["製品規格名称"]=="不足") | (df["製品規格名称"]=="従業員数")
        df = df[~delete_row]
        self.schedule_data = df
    
    def add_excel_function(self):
        try:
            wb = xl.load_workbook(filename=self.schedule_data_path)
            ws = wb[self.sheet_name]
        except Exception as e:
            raise Exception(f"Excelファイルを開けません\n{e}")
        
        last_row = ws.max_row
        ws.cell(row = last_row+1, column = 1).value = "欠勤"
        ws.cell(row = last_row+2, column = 1).value = "サンワ"
        ws.cell(row = last_row+3, column = 1).value = "合計"
        ws.cell(row = last_row+4, column = 1).value = "不足"
        ws.cell(row = last_row+6, column = 1).value = "従業員数"
        ws.cell(row = last_row+6, column = 2).value = "(入力してください)"
        
        for i, cell in enumerate(ws[last_row+3]):
            if i < 3:
                continue
            ws[cell.coordinate] = f"=SUM({cell.column_letter+str(cell.row-4)}:{cell.column_letter+str(cell.row-1)})"
        
        staff_num_cell = ws.cell(row = last_row+6, column = 2).coordinate
        for i, cell in enumerate(ws[last_row+4]):
            if i < 3:
                continue
            ws[cell.coordinate] = f"={staff_num_cell}-{cell.column_letter+str(cell.row-1)}"
                    
        wb.save(self.schedule_data_path)
    
    def calc(self):
        self.reset_calc()
        miss_datas = self.basecalc()
        self.adjust_cell_width()
        self.coloring()
        self.add_excel_function()
        
        return miss_datas


def test_usecase1():
    test_data = r"C:\Users\Owner\Downloads\学内アルバイト_ミニDX化\mini_DX_aporon\test_data_Sheet1_schedule.xlsx"
    c = Calculation(test_data)
    miss_datas = c.calc()
    if miss_datas:
        for miss_data in miss_datas:
            print(f"製品名：{miss_data[0]}について、\n工程名：{miss_data[1]}が見つかりませんでした。")
