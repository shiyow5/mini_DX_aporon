import openpyxl
import pandas as pd
import unicodedata
import datetime


def get_deadline(excel_path, sheet_name = None):
    if sheet_name:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
    else:
        df = pd.read_excel(excel_path)
    
    required_columns = {"製品規格名称", "名称2"}
    if not required_columns.issubset(df.columns):
        raise ValueError(f"以下の列が見つかりません:\n{required_columns - set(df.columns)}\n\n文字に間違いがないか確認してください。")
    
    delete_row = pd.isna(df["製品規格名称"]) | pd.isna(df["名称2"])
    df = df[~delete_row]
    
    df.columns = [pd.to_datetime(col, origin="1899-12-30", unit="D").date() if isinstance(col, (int, float)) else col for col in df.columns]
    df.columns = [col.date() if isinstance(col, datetime.datetime) else col for col in df.columns]
    
    return df


"""
def get_deadline(excel_path, sheet_name = None, target_rgb=(0, 176, 240)):
    # openpyxl を使用してワークブックを読み込む
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # ヘッダーを取得
    header = [cell.value for cell in ws[1]]

    # 必要な列が存在するか確認
    required_columns = {"製品規格名称", "名称2"}
    if not required_columns.issubset(header):
        missing_columns = required_columns - set(header)
        raise ValueError(f"以下の列が見つかりません:\n{missing_columns}\n\n文字に間違いがないか確認してください。")

    # 色付きセルを含む行を抽出
    data = []
    for row in ws.iter_rows(min_row=2):
        row_data = [cell.value for cell in row]
        has_target_color = False
        for cell in row:
            fill = cell.fill
            if fill and fill.fgColor and fill.fgColor.type == 'rgb':
                cell_rgb = fill.fgColor.rgb
                # RGB値を比較
                if cell_rgb == 'FF{:02X}{:02X}{:02X}'.format(*target_rgb):
                    has_target_color = True
                    break
        if has_target_color:
            data.append(row_data)

    # データフレームを作成
    df = pd.DataFrame(data, columns=header)
    
    delete_row = pd.isna(df["製品規格名称"]) | pd.isna(df["名称2"])
    df = df[~delete_row]
    
    df.columns = [pd.to_datetime(col, origin="1899-12-30", unit="D").date() if isinstance(col, (int, float)) else col for col in df.columns]
    df.columns = [col.date() if isinstance(col, datetime.datetime) else col for col in df.columns]
    
    return df
"""


def convert_half_to_full_katakana(text):
    """
    半角カタカナを全角カタカナに変換する関数。

    Args:
        text (str): 入力文字列。

    Returns:
        str: 半角カタカナが全角カタカナに変換された文字列。
    """
    # 全角への変換 (NFKC正規化)
    normalized_text = unicodedata.normalize('NFKC', text)
    return normalized_text


def clean_name2(text: str):
    ignores = ["トレー並べ", "計量", "梱包", "ダストヘラー"]
    
    text = convert_half_to_full_katakana(text)
    word_list = [clean_name(word) for word in text.split("・")]
    for ignore in ignores:
        if ignore in word_list:
            word_list.remove(ignore)
    
    return "・".join(word_list)
        

def clean_name(text: str):
    text = convert_half_to_full_katakana(text)
    text = text.replace("釦", "スパウト")
    text = text.replace("ボタン", "スパウト")
    text = text.replace("付け", "嵌め")
    text = text.replace("トップキャップ", "キャップ")
    #text = text.replace("ネジ嵌合・スパウト嵌め", "?")
    
    return text
